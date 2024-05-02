from collections import namedtuple
import logging
import datetime
import pickle
import struct
from multiprocessing.pool import ThreadPool

import psycopg2

from config import CUSTOM_SCHEMAS, LIST_LOAD, POSTGIS_TYPES_MAPPING, POSTGRES_DB, POSTGRES_PASSWORD, POSTGRES_PORT, POSTGRES_SERVER, POSTGRES_USER, engine_Db_no_async, APP, CHUNCKSIZE
import os

from utility import find_specTable, change_column_types

logger = logging.getLogger(APP)
from simpledbf import Dbf5


def get_now():
    return datetime.datetime.now()


class ImporterException(Exception):
    def __init__(self, message,e):
        self.e = e
        self.message = message
        super().__init__(f"Exception : {self.e} - {self.message}")

def clean_group_id(group_id,exists=False):
    return clean_column(group_id)

def clean_column(group_id,exists=False):
    if group_id is not None:
        group_id = group_id.lower() if group_id is not None else group_id
        if group_id in CUSTOM_SCHEMAS:
            group_id =f"__{group_id}__"
        
        for special in ["-","&","|","+",":",";","%","&","*"," "]:
            group_id =group_id.replace(special,"_")
        group_id = group_id.strip()
        for special in [f"{n}" for n in range(9)]:
            if group_id.startswith(special):
                group_id="_"+group_id+"_"
    return group_id

class Dbf_wrapper(Dbf5):

    def _get_recs(self, chunk=None):
        '''Generator that returns individual records.

        Parameters
        ----------
        chunk : int, optional
            Number of records to return as a single chunk. Default 'None',
            which uses all records.
        '''
        if chunk == None:
            chunk = self.numrec

        for i in range(chunk):
            # Extract a single record
            record = struct.unpack(self.fmt, self.f.read(self.fmtsiz))
            # If delete byte is not a space, record was deleted so skip
            if record[0] != b' ':
                continue

                # Save the column types for later
            self._dtypes = {}
            result = []
            for idx, value in enumerate(record):
                name, typ, size = self.fields[idx]
                if name == 'DeletionFlag':
                    continue

                # String (character) types, remove excess white space
                if typ == "C":
                    if name not in self._dtypes:
                        self._dtypes[name] = "str"
                    value = value.strip()
                    # Convert empty strings to NaN
                    if value == b'':
                        value = self._na
                    else:
                        value = value.decode(self._enc)
                        # Escape quoted characters
                        if self._esc:
                            value = value.replace('"', self._esc + '"')

                # Numeric type. Stored as string
                elif typ == "N":
                    # A decimal should indicate a float
                    if b'.' in value:
                        if name not in self._dtypes:
                            self._dtypes[name] = "float"
                        value = float(value)
                    # No decimal, probably an integer, but if that fails,
                    # probably NaN
                    else:
                        try:
                            value = int(value)
                            if name not in self._dtypes:
                                self._dtypes[name] = "int"
                        except:
                            # I changed this for SQL->Pandas conversion
                            # Otherwise floats were not showing up correctly
                            value = float('nan')

                # Date stores as string "YYYYMMDD", convert to datetime
                elif typ == 'D':
                    try:
                        y, m, d = int(value[:4]), int(value[4:6]), \
                            int(value[6:8])
                        if name not in self._dtypes:
                            self._dtypes[name] = "date"
                    except:
                        value = self._na
                    else:
                        if y==0 or m==0 or d==0:
                            value = None
                        else:
                            value = datetime.date(y, m, d)

                # Booleans can have multiple entry values
                elif typ == 'L':
                    if name not in self._dtypes:
                        self._dtypes[name] = "bool"
                    if value in b'TyTt':
                        value = True
                    elif value in b'NnFf':
                        value = False
                    # '?' indicates an empty value, convert this to NaN
                    else:
                        value = self._na

                # Floating points are also stored as strings.
                elif typ == 'F':
                    if name not in self._dtypes:
                        self._dtypes[name] = "float"
                    try:
                        value = float(value)
                    except:
                        value = float('nan')

                else:
                    err = 'Column type "{}" not yet supported.'
                    raise ValueError(err.format(value))

                result.append(value)
            yield result


def shapeFile2Postgis(validation_id,map_files,map_tables_edited,group_id,conn_str,schema=None,
                      engine=engine_Db_no_async, srid=None,load_type="append",multithread=True):
    from importerLayers import publish_layers
    try:
        if not schema:
            schema=group_id
        map_results = {}
        map_filtered= {}
        for column in map_tables_edited.data:
            if column.filename not in map_filtered:
                map_filtered[column.filename]=[]
            if column.column not in [column.filename] and column.importing:
                map_filtered[column.filename].append(column)
            
        list_dbf=[k for k, file_name in map_files.items() if file_name.endswith('.dbf') and file_name in map_filtered]
        list_csv=[k for k, file_name in map_files.items() if file_name.endswith('.csv') and file_name in map_filtered]
        list_shp=[k for k, file_name in map_files.items() if file_name.endswith('.shp') and file_name in map_filtered]
        list_excel=[k for k, file_name in map_files.items() if file_name.endswith('.xls') or file_name.endswith('.xlsx')]
        
        map_exists = {} 
        for table in list_dbf:
            pickle_file =os.path.join("data",f"{table}.pickle")
            if os.path.exists(pickle_file):
                map_exists[table]=pickle_file
        
        list_dbf=[k for k, file_name in map_files.items() if file_name.endswith('.dbf') and file_name in map_filtered and k not in map_exists]
        list_csv=[k for k, file_name in map_files.items() if file_name.endswith('.csv') and file_name in map_filtered and k not in map_exists]
        list_shp=[k for k, file_name in map_files.items() if file_name.endswith('.shp') and file_name in map_filtered and k not in map_exists]
        list_excel=[k for k, file_name in map_files.items() if file_name.endswith('.xls') or file_name.endswith('.xlsx') and file_name in map_filtered]
        logger.info(f"validation_id:{validation_id}, dbf:{len(list_dbf)},shp:{len(list_shp)},csv:{len(list_csv)},excel:{len(list_excel)}")
                
        start_time_tot = get_now()
        if multithread: 
            logger.info(f"Caricamento parallelo, srid={srid}, load_type={load_type}")
            num_thread =int(len(map_files)/2)+1
            tasks=[]
            pool = ThreadPool(num_thread)
            for k in list_dbf:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                task = pool.apply_async(load_dbf_to_postgis, args=(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid))
                tasks.append(task)
            for k in list_csv:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                task = pool.apply_async(load_csv_to_postgis, args=(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid))
                tasks.append(task)
            for k in list_excel:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                task = pool.apply_async(load_excel_to_postgis, args=(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid))
                tasks.append(task)
            for k in list_shp:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                task = pool.apply_async(load_shapefile_to_postgis, args=(file_name, map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid))
                tasks.append(task)
            pool.close()
            pool.join()

            results = [r.get() for r in tasks if r.get() is not None]
        else:
            results = []
            logger.info(f"Caricamento sincrono, srid={srid}, load_type={load_type}")
            for k in list_dbf:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                res = load_dbf_to_postgis(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid)
                results.append(res)
            for k in list_csv:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                res = load_csv_to_postgis(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid)
                results.append(res)
            for k in list_excel:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                res = load_excel_to_postgis(file_name,map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid)
                results.append(res)
            for k in list_shp:
                file_name = map_files[k]
                table_name = k  # Usa il nome del file shape come nome della tabella
                res = load_shapefile_to_postgis(file_name, map_tables_edited, table_name,conn_str,schema,engine,group_id,load_type,srid)
                results.append(res)
            
        map_errors = {}
        for res,table_name,elapsed,have_geom in results:
            file_name = map_files[table_name]
            if res:
                map_results[file_name]=table_name
                if have_geom and table_name not in list_shp:
                    list_shp.append(table_name)
            else:
                if "Error" not in map_results:
                    map_results["Error"]=[]
                map_errors[file_name]=file_name
                map_results["Error"].append(file_name)
            if "Stats" not in map_results:
                map_results["Stats"]={}
            map_results["Stats"][table_name]=f"{file_name} elapsed:{elapsed}"
        if "Stats" not in map_results:
            map_results["Stats"]={}
        elapsed=(get_now() - start_time_tot).total_seconds()
        map_results["Stats"]["TOTAL"]=f"Total time elapsed:{elapsed}"
        logger.info(f"terminati shp:{len(list_shp)},Total time elapsed:{elapsed}")
        layers = [table for table in list_shp if table not in map_errors]
        publish_layers(group_id,layers=layers,with_view=True)
        if len(map_exists)>0:
            map_results["source_duplicated"]=map_exists
        return map_results
    except Exception as e:
        logger.error(f"Error shapeFile2Postgis  validation_id:{validation_id} error:{e}", exc_info=True)
        print(e)


def load_dbf(shapefile_path, table_name,group_id=None,srid=None):
    try:
        start_time = get_now()
        dbf = Dbf_wrapper(shapefile_path)
        df = dbf.to_dataframe()
        df = df.rename(columns=str.lower)
        df = df.rename(columns=clean_column)
        
        if group_id:
            df['group_id']=group_id
        map_create, columns, _, columns_list = get_columns_shapefile(shapefile_path,table_name,df,srid)
        elapsed = (get_now() - start_time).total_seconds()
        return map_create, columns, _, columns_list,df,elapsed
    except Exception as e:
        raise ImporterException(f"Error {e} load {shapefile_path}",e)

def load_csv(shapefile_path, table_name,group_id=None,srid=None):
    import pandas as pd
    import geopandas as gpd
    from shapely import wkt
    import chardet
    start_time = get_now()
    df=None
    encoding = None
    try:
        csv = None
        with open(shapefile_path, 'rb') as f:
            r = f.readline()
            encoding = chardet.detect(r)
            encoding = namedtuple("Encoding",list(encoding.keys()))(**encoding)
            row=r.decode(encoding.encoding) 
            if encoding.encoding!="utf_8":
                csv = row.replace("\r","").replace("\n","").replace("'","")+"\n" 
            for sep in[";","\t"]: 
                if len(row.split(sep))>1: 
                    break
            for delimiter in["\r\n","\n","\r"]: 
                if len(row.split(delimiter))>1: 
                    break
            if csv:
                for line in f:
                    csv+=line.decode(encoding.encoding,errors="ignore").replace("\r","").replace("\n","").replace("'","")+"\n" 
        if csv:
            with open(shapefile_path, 'wb') as f:
                f.write(csv.encode())
                    
            df = pd.read_csv(shapefile_path,sep=sep,on_bad_lines="skip",lineterminator='\n')
        else:
            df = pd.read_csv(shapefile_path)
        
        df = df.rename(columns=str.lower)
        df = df.rename(columns=clean_column)

        if group_id:
            df['group_id']=group_id
        columns =list(df.columns)
            
        if "latitudine" in columns and "longitudine" in columns:
            srid = 4326
            df = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.longitudine, df.latitudine), crs='epsg:4326')
        elif "latitude" in columns and "longitude" in columns:
            srid = 4326
            df = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.longitude, df.latitude), crs='epsg:4326')
        elif "geometry" in columns:
            srid = 4326
            df["geometry"] = df["geometry"].apply(wkt.loads)
            df = gpd.GeoDataFrame(df, crs='epsg:4326')
        map_create, columns, _, columns_list = get_columns_shapefile(shapefile_path,table_name,df,srid)
        elapsed = (get_now() - start_time).total_seconds()
        return map_create, columns, srid, columns_list,df,elapsed
    except Exception as e:
        raise ImporterException(f"Error {e} load {shapefile_path} excpected utf8 encoding:{encoding}",e)
        
def load_excel(shapefile_path, table_name,group_id=None,srid=None):
    import pandas as pd
    import geopandas as gpd
    from shapely import wkt
    from openpyxl import load_workbook
    try:
        wb = load_workbook(filename = shapefile_path, data_only = True)
        sheet_names = wb.sheetnames
        map_total ={}
        for sheet_name in sheet_names:
            df = pd.read_excel(shapefile_path,sheet_name)
            df = df.rename(columns=str.lower)
            df = df.rename(columns=clean_column)
            columns =list(df.columns)
            if "latitudine" in columns and "longitudine" in columns:
                srid = 4326
                df = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.longitudine, df.latitudine), crs='epsg:4326')
            elif "latitude" in columns and "longitude" in columns:
                srid = 4326
                df = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.longitude, df.latitude), crs='epsg:4326')
            elif "geometry" in columns:
                srid = 4326
                df["geometry"] = df["geometry"].apply(wkt.loads)
                df = gpd.GeoDataFrame(df, crs='epsg:4326')
            if group_id:
                df['group_id']=group_id
            map_create, columns, _, columns_list = get_columns_shapefile(shapefile_path,table_name,df,srid)
            info ={"map_create":map_create, "columns":columns,"columns_list":columns_list,"table_name":table_name,"df":df}
            map_total[sheet_name]=namedtuple("Table",list(info))(**info)
        return map_total
    except Exception as e:
        raise ImporterException(f"Error {e} load {shapefile_path}",e)
    
def load_csv_to_postgis(shapefile_path,map_tables_edited,table_name,conn_str,schema,engine,group_id,load_type,srid_validation):
    try:
        #import pandas as pd
        start_time = get_now()
        map_create, columns, _, columns_list,df,elapsed=load_csv(shapefile_path,table_name,group_id,srid_validation)
        json_types = find_specTable(map_tables_edited,table_name)
        df = change_column_types(df, json_types)
        res = load_df_to_postgres(load_type,conn_str,schema,table_name,columns,df,columns_list,engine)
        elapsed=(get_now() - start_time).total_seconds()
        logger.info(f"caricato {table_name},map_create:{map_create},elapsed:{elapsed}")
        return res,table_name,elapsed,hasattr(df,"to_postgis")
    except Exception as e:
        logger.error(f"Error load_csv_to_postgis table {table_name}: {e}",stack_info=True)
        elapsed=(get_now() - start_time).total_seconds()
        return False,table_name,elapsed,hasattr(df,"to_postgis")

def load_excel_to_postgis(shapefile_path,map_tables_edited,table_name,conn_str,schema,engine,group_id,load_type,srid_validation):
    try:
        #import pandas as pd
        start_time = get_now()
        map_total=load_excel(shapefile_path,table_name,group_id,srid_validation)
        for sheet_name,info in map_total.items():
            df = info.df
            columns = info.columns
            columns_list = info.columns_list
            map_create = info.map_create
            json_types=find_specTable(map_tables_edited,sheet_name)
            df = change_column_types(df, json_types)
            res = load_df_to_postgres(load_type,conn_str,schema,table_name,columns,df,columns_list,engine)
            elapsed=(get_now() - start_time).total_seconds()
            logger.info(f"caricato {table_name},map_create:{map_create},elapsed:{elapsed}")
        elapsed=(get_now() - start_time).total_seconds()
        return res,table_name,elapsed,hasattr(df,"to_postgis")
    except Exception as e:
        logger.error(f"Error load_excel_to_postgis table {table_name}: {e}",stack_info=True)
        elapsed=(get_now() - start_time).total_seconds()
        return False,table_name,elapsed,hasattr(df,"to_postgis")

def load_dbf_to_postgis(shapefile_path,map_tables_edited,table_name,conn_str,schema,engine,group_id,load_type,srid_validation):
    try:
        #import pandas as pd
        start_time = get_now()
        map_create, columns, _, columns_list,df,elapsed=load_dbf(shapefile_path,table_name,group_id,srid_validation)
        json_types=find_specTable(map_tables_edited,table_name)
        df = change_column_types(df, json_types)
        res = load_df_to_postgres(load_type,conn_str,schema,table_name,columns,df,columns_list,engine)
        elapsed=(get_now() - start_time).total_seconds()
        logger.info(f"caricato {table_name},map_create:{map_create},elapsed:{elapsed}")
        return res,table_name,elapsed,False
    except Exception as e:
        logger.error(f"Error load_dbf_to_postgis table {table_name}: {e}",stack_info=True)
        elapsed=(get_now() - start_time).total_seconds()
        return False,table_name,elapsed,False

def get_map_files(validation_id,conn_str_db):
    map_files={}
    with psycopg2.connect(conn_str_db) as conn:
        with conn.cursor() as cur:
            sql=f'SELECT "USERFILE" FROM geo_labs.richieste_upload where "ID_SHAPE"={validation_id};'
            cur.execute(sql)
            res = cur.fetchone()
            map_files =  {os.path.split(file_name)[-1][:-4].lower():file_name for file_name in res[0]}
    return map_files


def load_df_to_postgres(load_type,conn_str,schema,table_name,columns,df,columns_list,engine):
    sql =f"select count(*) from pg_tables where schemaname='{schema}' and tablename='{table_name}'"        
    with psycopg2.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            r = cur.fetchone()
            if r and r[0]==0 and load_type!=LIST_LOAD[1]:
                load_type = LIST_LOAD[1]
                logger.warning(f"Table  {schema}.{table_name} not found forced strict")
               
               
    if load_type==LIST_LOAD[1]:  #caso in cui si ricrea la tabella
        #columns=columns.replace("GEOMETRY","geometry")
        create_table_query = f"CREATE TABLE IF NOT EXISTS {schema}.{table_name} ({columns}"
        with psycopg2.connect(conn_str) as conn:
            conn.set_isolation_level(0)
            # Creazione della tabella
            with conn.cursor() as cur:
                #DROP TABELLE CON LO STESSO NOME IN FOREIGN TABLE
                try:
                    sql_drop_foreign_table_if_exists=f"DROP FOREIGN TABLE IF EXISTS {schema}.{table_name} CASCADE"
                    cur.execute(sql_drop_foreign_table_if_exists)
                    conn.commit()
                except:
                    conn.rollback()
                    logger.warning(f"No Foreign Table  {schema}.{table_name}")
                try:
                    sql_drop_table_if_exists = f"DROP TABLE IF EXISTS {schema}.{table_name} CASCADE"
                    cur.execute(sql_drop_table_if_exists)
                    conn.commit()
                except:
                    conn.rollback()
                    logger.warning(f"No Table  {schema}.{table_name}",stack_info=True)

                #if len(keys)>0:
                #     keys.append("group_id")
                #     primary_key="("+",".join(keys)+")"
                #     create_table_query+=f" ,CONSTRAINT {table_name}_pk PRIMARY KEY {primary_key})"
                # else:
                #     create_table_query+=")"

                create_table_query+=")"
                cur.execute(create_table_query)
                conn.commit()
                # Caricamento dei dati nella tabella
                if hasattr(df,"to_postgis"):
                    df.to_postgis(table_name, engine, if_exists="append", index=False, schema=schema,chunksize=CHUNCKSIZE)
                else:
                    df.to_sql(table_name, engine, if_exists="append", index=False, schema=schema,chunksize=CHUNCKSIZE)
                pickle_file = os.path.join("data",f"{table_name}.pickle")
                with open(pickle_file, 'wb') as f:
                    pickle.dump(df,f)
                engine.dispose()
                return True
    else:
        table_name_temp =f"temp_{table_name}"
        create_table_query = f"CREATE TABLE IF NOT EXISTS {schema}.{table_name_temp} ({columns}"
        if hasattr(df,"to_postgis"):
            df.to_postgis(table_name_temp, engine, if_exists=load_type, index=False, schema=schema,chunksize=CHUNCKSIZE)
        else:
            df.to_sql(table_name_temp, engine, if_exists=load_type, index=False, schema=schema,chunksize=CHUNCKSIZE)
        engine.dispose()
        join_columns=','.join(columns_list)
        join_temp_columns=','.join([f"{table_name_temp}.{k}" for k in columns_list])
        join_excluded=','.join([f"{k} = EXCLUDED.{k}" for k in columns_list if k not in ["id"]])

        upsert_query = f"""
            INSERT INTO {table_name} ({join_columns})
            SELECT {join_temp_columns}
            FROM {table_name_temp}
            ON CONFLICT ({table_name}_pk) DO UPDATE
            SET {join_excluded}
        """
        with engine.connect() as conn:
            conn.execute(upsert_query)

        with psycopg2.connect(conn_str) as conn:
            conn.set_isolation_level(0)
            # Creazione della tabella
            with conn.cursor() as cur:
                try:
                    sql_drop_table_if_exists = f"DROP TABLE IF EXISTS {schema}.{table_name_temp}"
                    cur.execute(sql_drop_table_if_exists)
                    conn.commit()
                except:
                    conn.rollback()
                    logger.warning(f"No Table  {schema}.temp_{table_name}",stack_info=True)
        return True

def get_columns_shapefile(shapefile_path,table_name,gdf, srid):
    columns_list = list(gdf.columns)
    columns_list.append("group_id")
    if hasattr(gdf,"crs"):
        crs_string = str(gdf.crs)
        geometry_type = gdf.geometry.geom_type.unique()[0]
        # verifica se ci sta la coordinata Z:
        has_z_coordinate = gdf.geometry.has_z.all()
        if has_z_coordinate:
            geometry_type += "Z"
        logger.info(f"carico {shapefile_path} geometry_type:{geometry_type}")

        if "AUTHORITY" in crs_string  and srid is None:
            srid = int(crs_string.split("EPSG")[-1].replace("\"", "").replace("]", "").replace(",", ""))
        representation = f"geometry({geometry_type},{srid})"
    else:
        representation = f"geometry"
    
    gdf_types_list = []

    for dtype in gdf.dtypes:
        if dtype == "geometry":
            tipo = representation
        else:
            tipo = POSTGIS_TYPES_MAPPING[str(dtype)]
        gdf_types_list.append(tipo)
    gdf_types_list.append("text")  # group_id
    # se esiste un campo di tipo geometry, aggiungere un campo "geometry" tipo geometry("tipo senza Z",25832)
    #for ele in gdf_types_list:
    #    if ele.startswith("geometry") and "Z," in ele.upper():
    #        comma_index = ele.find(",")
    #        if comma_index != -1:
    #            number_str = ele[comma_index + 1:-1]
    #            ele_new = ele.replace("Z", "").replace("z", "")#.replace(number_str, "25832")
    #            #columns_list.append("geometry_2D")
    #            gdf_types_list.append(ele_new)
    map_create = {clean_column(key): value for key, value in zip(columns_list, gdf_types_list)}
    map_create_date = {}

    map_create.update(map_create_date)
    columns = ", ".join([f"{clean_column(column)} {dtype}" for column, dtype in map_create.items()])
    return map_create,columns,srid,columns_list

def load_shapefile_to_postgis(shapefile_path,map_tables_edited,table_name,conn_str,schema,engine,group_id,load_type,srid_validation):
    map_create={}
    try:
        start_time = get_now()
        resp, columns, gdf, columns_list, start_time, elapsed,map_create = load_shapefile(shapefile_path, table_name, group_id, srid_validation)
        json_types=find_specTable(map_tables_edited,table_name)
        gdf=change_column_types(gdf, json_types)
        if not resp:
            return False,table_name,elapsed,False
        res = load_df_to_postgres(load_type,conn_str,schema,table_name,columns,gdf,columns_list,engine)
        return res,table_name,elapsed,True
    except Exception as e:
        elapsed=(get_now() - start_time).total_seconds()
        logger.error(f"Error creating table {table_name}: {e},map_create:{map_create}")
        return False,table_name,elapsed,True

def load_shapefile(shapefile_path,table_name,group_id,srid):
    import geopandas as gpd
    start_time = get_now()
    gdf = gpd.read_file(shapefile_path)
    gdf = gdf.rename(columns=str.lower)
    if srid is not None:
        gdf.set_crs(f"EPSG:{srid}", allow_override=True,inplace=True)
    else:
        srid = gdf.crs.to_epsg()
    if group_id:
        gdf['group_id']=group_id
    geometry_type = gdf.geometry.geom_type.unique()[0]
    if geometry_type is None:
        elapsed=(get_now() - start_time).total_seconds()
        logger.error(f"{shapefile_path} non la geometria valorizzata"  )
        return False, table_name, elapsed
    map_create, columns, srid, columns_list = get_columns_shapefile(shapefile_path,table_name,gdf,srid)
    elapsed = (get_now() - start_time).total_seconds()
    return True, columns, gdf, columns_list, start_time, elapsed,map_create

def create_schema(group_id,conn_str=f"host='{POSTGRES_SERVER}' port='{POSTGRES_PORT}' dbname='{POSTGRES_DB}' user='{POSTGRES_USER}' password='{POSTGRES_PASSWORD}'"):
    sql =f"select count(*) from information_schema.schemata where schema_name='{group_id}'"
    db = psycopg2.connect(conn_str)
    logger.info(f"start {conn_str}")
    res=0
    with db.cursor() as cursor:
        cursor.execute(sql)
        results = cursor.fetchall()
        if len(results)>0:
            res=results[0][0]
        if res<1:
            sql=f"CREATE SCHEMA IF NOT EXISTS {group_id} AUTHORIZATION postgres;\n"
            sql+=f"COMMENT ON SCHEMA {group_id} IS 'standard {group_id} schema';\n"
            sql+=f"GRANT ALL ON SCHEMA {group_id} TO PUBLIC;\n"
            sql+=f"GRANT ALL ON SCHEMA {group_id} TO postgres;\n"
            cursor.execute(sql)
            db.commit()
        logger.info(f"Created schema {group_id} on Db:{conn_str}")